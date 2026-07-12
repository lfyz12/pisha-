import io
from dataclasses import dataclass, field
from typing import Any

import openpyxl


@dataclass
class ExcelStudentRaw:
    group_name: str
    full_name: str
    total_score: float
    average_score: float
    attendance: list[float] = field(default_factory=list)
    science_activity: dict[str, float] = field(default_factory=dict)
    project_activity: dict[str, float] = field(default_factory=dict)
    extracurricular: dict[str, float] = field(default_factory=dict)


@dataclass
class ExcelEvent:
    id: int
    name: str
    category: str
    date: str
    level: str
    status: str
    points: float


@dataclass
class ParsedExcelData:
    students: list[ExcelStudentRaw]
    events: list[ExcelEvent]


def _to_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", ".").replace(" ", "").replace("\u00a0", "")
        cleaned = "".join(c for c in cleaned if c.isdigit() or c == ".")
        if not cleaned:
            return 0.0
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    return 0.0


def _fill_forward(row: list[Any]) -> list[Any]:
    result = []
    last = None
    for cell in row:
        if cell not in (None, ""):
            last = cell
        result.append(last)
    return result


HEADER_KEYWORDS = ["групп", "фио", "номер", "балл", "посещаем", "категори", "успеваем"]


def _is_meta_row(row: list[str], idx: int) -> bool:
    first4 = row[:4]
    all_empty = all(not c for c in first4)
    if not all_empty:
        return False
    rest_text = [c for c in row[4:] if c]
    has_meta = any(
        "количеств" in t or "учебн" in t or "семестр" in t for t in rest_text
    )
    return has_meta or (idx == 0 and bool(rest_text))


def _is_header_row(row: list[str]) -> bool:
    first4 = row[:4]
    return any(
        any(kw in c for kw in HEADER_KEYWORDS) for c in first4 if c
    )


def parse_rating_excel_buffer(buffer: bytes) -> ParsedExcelData:
    workbook = openpyxl.load_workbook(filename=io.BytesIO(buffer), data_only=True)
    main_sheet = workbook.worksheets[0]
    raw_rows = [
        [_to_string(cell.value).lower() for cell in row]
        for row in main_sheet.iter_rows()
    ]
    events = _parse_events_sheet(workbook)
    students = _parse_main_sheet(raw_rows)
    return ParsedExcelData(students=students, events=events)


def _parse_events_sheet(workbook: openpyxl.Workbook) -> list[ExcelEvent]:
    events_sheet = next(
        (
            sheet
            for sheet in workbook.worksheets
            if "перечень" in sheet.title.lower() or "мероприя" in sheet.title.lower()
        ),
        None,
    )
    if events_sheet is None:
        return []

    rows = list(events_sheet.iter_rows(values_only=True))
    events = []
    for idx, row in enumerate(rows):
        if not row or len(row) < 7:
            continue
        events.append(
            ExcelEvent(
                id=_to_number(row[0]) or idx + 1,
                name=_to_string(row[1]),
                category=_to_string(row[2]),
                date=_to_string(row[3]),
                level=_to_string(row[4]),
                status=_to_string(row[5]),
                points=_to_number(row[6]),
            )
        )
    return events


def _parse_main_sheet(raw_rows: list[list[str]]) -> list[ExcelStudentRaw]:
    if len(raw_rows) < 4:
        return []

    header_start_index = 0
    for i in range(min(10, len(raw_rows))):
        row = raw_rows[i]
        if not row:
            continue
        if _is_meta_row(row, i):
            continue
        if _is_header_row(row):
            header_start_index = i
            break

    header_rows: list[list[str]] = []
    data_start_index = header_start_index
    for i in range(header_start_index, min(header_start_index + 5, len(raw_rows))):
        row = raw_rows[i]
        if not row:
            continue
        first_cell = row[0]
        is_still_header = (
            first_cell == ""
            or any(kw in first_cell for kw in HEADER_KEYWORDS)
            or len(header_rows) < 3
        )
        if is_still_header:
            header_rows.append(row)
        else:
            data_start_index = i
            break

    if len(header_rows) < 2:
        return []

    max_cols = max(len(r) for r in raw_rows) if raw_rows else 5
    filled_headers = []
    for r in header_rows:
        padded = r + [None] * (max_cols - len(r))
        filled_headers.append(_fill_forward(padded))

    num_header_rows = len(filled_headers)
    categories: list[dict[str, Any] | None] = []

    for col in range(max_cols):
        top_label = _to_string(filled_headers[0][col])
        sub_labels = [
            _to_string(filled_headers[h][col])
            for h in range(1, num_header_rows)
        ]

        if "групп" in top_label or "номер" in top_label:
            categories.append({"type": "base", "sub_key": "group_name"})
        elif any(kw in top_label for kw in ["фио", "фам", "имя", "студент"]):
            categories.append({"type": "base", "sub_key": "full_name"})
        elif "итогов" in top_label or ("балл" in top_label and "средн" not in top_label):
            categories.append({"type": "base", "sub_key": "total_score"})
        elif "средн" in top_label or ("балл" in top_label and "успеваем" in top_label):
            categories.append({"type": "base", "sub_key": "average_score"})
        elif "посещаем" in top_label or "образов" in top_label:
            week_label = next((s for s in sub_labels if s.isdigit()), f"week-{col}")
            categories.append({"type": "attendance", "sub_key": week_label})
        elif "научн" in top_label:
            label = sub_labels[0] if sub_labels else f"science-{col}"
            categories.append({"type": "science", "sub_key": label})
        elif "проект" in top_label:
            label = sub_labels[0] if sub_labels else f"project-{col}"
            categories.append({"type": "project", "sub_key": label})
        elif "внеучеб" in top_label or "внеуч" in top_label:
            label = sub_labels[0] if sub_labels else f"extracurr-{col}"
            categories.append({"type": "extracurricular", "sub_key": label})
        else:
            categories.append(None)

    students: list[ExcelStudentRaw] = []
    for i in range(data_start_index, len(raw_rows)):
        row = raw_rows[i]
        if not row:
            continue
        first_cell = _to_string(row[0])
        has_numeric = any(
            isinstance(cell, (int, float)) or (isinstance(cell, str) and any(c.isdigit() for c in cell))
            for cell in row
        )
        if not first_cell or (not first_cell.replace("-", "").replace(" ", "").isalpha() and not has_numeric):
            continue

        student = ExcelStudentRaw(group_name="", full_name="", total_score=0.0, average_score=0.0)
        attendance_map: dict[str, float] = {}
        science_map: dict[str, float] = {}
        project_map: dict[str, float] = {}
        extracurr_map: dict[str, float] = {}

        for col in range(min(len(row), len(categories))):
            cat = categories[col]
            if cat is None:
                continue
            cell_val = row[col]
            cat_type = cat["type"]
            sub_key = cat["sub_key"]

            if cat_type == "base":
                if sub_key == "group_name":
                    student.group_name = _to_string(cell_val)
                elif sub_key == "full_name":
                    student.full_name = _to_string(cell_val)
                elif sub_key == "total_score":
                    student.total_score = _to_number(cell_val)
                elif sub_key == "average_score":
                    student.average_score = _to_number(cell_val)
            elif cat_type == "attendance":
                attendance_map[sub_key] = _to_number(cell_val)
            elif cat_type == "science":
                science_map[sub_key] = _to_number(cell_val)
            elif cat_type == "project":
                project_map[sub_key] = _to_number(cell_val)
            elif cat_type == "extracurricular":
                extracurr_map[sub_key] = _to_number(cell_val)

        def sort_key(k: str):
            try:
                return (0, int(k))
            except ValueError:
                return (1, k)

        student.attendance = [attendance_map[k] for k in sorted(attendance_map.keys(), key=sort_key)]
        student.science_activity = science_map
        student.project_activity = project_map
        student.extracurricular = extracurr_map

        if student.full_name:
            students.append(student)

    return students

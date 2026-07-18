import io

import openpyxl
from django.test import TestCase
from django.core.files.uploadedfile import SimpleUploadedFile

from import_export.parsers import (
    ExcelStudentRaw,
    has_multi_level_header,
    parse_flat_excel_buffer,
    parse_rating_excel_buffer,
)
from import_export.views import _validate_xlsx


class ParserSmokeTests(TestCase):
    databases = []

    def test_flat_parser_reads_simple_columns(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Студент", "Группа", "Курс", "Баллы (учеба)", "Баллы (активность)", "Общий балл"])
        sheet.append(["Иван Петров", "ИС-101", 2, 80, 20, 100])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        result = parse_flat_excel_buffer(buffer.getvalue())

        self.assertEqual(len(result), 1)
        student = result[0]
        self.assertIsInstance(student, ExcelStudentRaw)
        self.assertEqual(student.full_name, "Иван Петров")
        self.assertEqual(student.group_name, "ИС-101")

    def test_multi_level_parser_detects_header_and_data(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["", "", "", "", "Категория"])
        sheet.append(["Группа", "ФИО", "Итоговый балл", "Средний балл", "Посещаемость"])
        sheet.append(["", "", "", "", "1"])
        sheet.append(["ИС-101", "Иван Петров", 100, 80, 90])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        parsed = parse_rating_excel_buffer(buffer.getvalue())

        self.assertEqual(len(parsed.students), 1)
        student = parsed.students[0]
        self.assertEqual(student.full_name, "Иван Петров")
        self.assertEqual(student.total_score, 100)

    def test_has_multi_level_header_detects_keyword(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Категория", "Группа", "ФИО"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        result = has_multi_level_header(buffer.getvalue())

        self.assertTrue(result)

    def test_auto_parser_falls_back_to_multi_level(self):
        from import_export.parsers import parse_excel_auto

        # Same shape as the real 1C export: first row has no keywords
        # for the header sniffer, actual header lives in row 3.
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["", "", "", "", "Количество учебных часов"])
        sheet.append(["", "", "", "", "1", "2"])
        sheet.append(["Номер \nгруппы", "ФИО", "Итоговый \nбалл", "Средний балл", "Посещаемость"])
        sheet.append(["ИС-101", "Иван Петров", 100, 80, 90])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        parsed = parse_excel_auto(buffer.getvalue())

        self.assertEqual(len(parsed.students), 1)
        self.assertEqual(parsed.students[0].full_name, "Иван Петров")

    def test_rejects_non_xlsx_upload(self):
        upload = SimpleUploadedFile("rating.xls", b"not-an-xlsx")
        with self.assertRaisesMessage(ValueError, "Only .xlsx"):
            _validate_xlsx(upload, b"not-an-xlsx", "auto")

    def _xlsx_with_broken_stylesheet(self, rows):
        """Build a valid xlsx, then corrupt xl/styles.xml (1C-style export)."""
        import zipfile

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in rows:
            sheet.append(row)
        # Distinct cell formats => cells reference non-zero style ids,
        # which the repaired stylesheet must still cover.
        sheet["A1"].font = openpyxl.styles.Font(bold=True)
        sheet["B1"].fill = openpyxl.styles.PatternFill(
            fill_type="solid", start_color="FFFF00", end_color="FFFF00"
        )
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        broken = io.BytesIO()
        with zipfile.ZipFile(buffer) as source, zipfile.ZipFile(
            broken, "w", zipfile.ZIP_DEFLATED
        ) as out:
            for entry in source.infolist():
                data = source.read(entry.filename)
                if entry.filename == "xl/styles.xml":
                    data = b"<styleSheet><broken & invalid"
                out.writestr(entry, data)
        return broken.getvalue()

    def test_flat_parser_tolerates_broken_stylesheet(self):
        import warnings

        buffer = self._xlsx_with_broken_stylesheet(
            [
                ["Студент", "Группа", "Курс", "Баллы (учеба)", "Баллы (активность)", "Общий балл"],
                ["Иван Петров", "ИС-101", 2, 80, 20, 100],
            ]
        )

        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            result = parse_flat_excel_buffer(buffer)

        self.assertEqual(len(result), 1)
        self.assertEqual(result[0].full_name, "Иван Петров")
        self.assertFalse(
            [w for w in caught if "default style" in str(w.message)],
            "repaired stylesheet should not trigger openpyxl warnings",
        )

    def test_rating_parser_tolerates_broken_stylesheet(self):
        buffer = self._xlsx_with_broken_stylesheet(
            [
                ["", "", "", "", "Категория"],
                ["Группа", "ФИО", "Итоговый балл", "Средний балл", "Посещаемость"],
                ["", "", "", "", "1"],
                ["ИС-101", "Иван Петров", 100, 80, 90],
            ]
        )

        parsed = parse_rating_excel_buffer(buffer)

        self.assertEqual(len(parsed.students), 1)
        self.assertEqual(parsed.students[0].full_name, "Иван Петров")
